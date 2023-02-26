from .detect_human import look_for_human
from .telegram_utils import sight_bot
import pandas as pd 
import time
import os 
import xlwt
from openpyxl import load_workbook
from xlwt import Workbook
import json
import pandas as pd
from django.shortcuts import redirect
from django.contrib.auth import logout
class colors:
    red = '\033[91m'
    green = '\33[92m'
    end = '\033[0m'


def format_time(seconds): 
    seconds = seconds % (24 * 3600) 
    hour = seconds // 3600
    seconds %= 3600
    minutes = seconds // 60
    seconds %= 60
      
    return "%d:%02d:%02d" % (hour, minutes, seconds)


def start():
    f = open('usr_data.json','r')
    j = json.loads(f.read())
    student_name = j['fullname']
    student_roll_number = j['username']
    bot = sight_bot()
    # message = "Name = " + student_name +  " \nID  = " + student_roll_number + "\nattention status : " + str(foo)
    intro_message = student_name + " `{" + "ID: "+ student_roll_number +  "}`" + " just joined the class"
    resp = bot.send_message(intro_message)
    print("Started monitoring, press Ctrl+C to end :)")
    class_start_time = time.time()
    total_uptime = 0
    while True:
        if look_for_human() == True:
                bot.send_message('Recognised student :)')
                start_time = time.time()
                while True:
                    h = look_for_human()
                    if h == False:
                        end_time = time.time()
                        # print("uptime was :", end_time - start_time)
                        total_uptime += end_time - start_time
                        bot.send_message('Student not found :(')
                        break
                    else:
                        pass    
        else:
            class_end_time = time.time()
            downtime = class_end_time - class_start_time - total_uptime
            if downtime < 15:
                continue
            end_message = student_name + " `{" + "ID: "+ student_roll_number +  "}`" + " left the class \n`Uptime: " + str(format_time(total_uptime)) + "` \n`Downtime: " + str(format_time(downtime)) + "`"
            resp = bot.send_message(end_message)
            #writer = pd.ExcelWriter('attention.xlsx', engine='openpyxl')
            """writer.book = load_workbook('attention.xlsx')
            writer.sheets = dict((ws.title, ws) for ws in writer.book.worksheets)
            reader = pd.read_excel(r'attention.xlsx')
            df = pd.DataFrame({'Username': [student_roll_number], 
                                'Uptime': [format_time(total_uptime)],
                                'Downtime': [format_time(downtime)]})
            df.to_excel(writer, index=False, header=False,  startrow=len(reader)+1)
            writer.close()"""
            '''wb = Workbook()
            sheet1 = wb.add_sheet('Attention Report')
            style = xlwt.easyxf('font: bold 1')
            sheet1.write(0, 0, 'Username', style)
            sheet1.write(0, 1, 'Uptime', style)
            sheet1.write(0, 2, 'Downtime', style)
            sheet1.write(1, 0, student_roll_number)
            sheet1.write(1, 1, format_time(total_uptime))
            sheet1.write(1, 2, format_time(downtime))
            wb.save('attention.xls')'''
            print("\nClass ended \nUptime: " , format_time(total_uptime), "\nDowntime: ", format_time(downtime))
            #logout(request)
            #redirect('login')
            break
